#!/usr/bin/env python3
"""ALKIS – GeoAS – GEOgraf Assistent 1.2.0
Portable-GUI/CLI: OpenData ZIP/XML -> GeoAS-Link -> XLSX -> GEOgraf-NAS.
Bestaetigte Regeln: 1100, 1301, 1302, 2101, Rechtsgemeinschaften.
Fehlende OpenData-Flurstuecke in GeoAS fuehren zu GELB, nicht zum Abbruch.
"""
from __future__ import annotations
import argparse, hashlib, html, json, os, re, shutil, sys, tempfile, traceback, webbrowser, zipfile
from collections import Counter, defaultdict
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation
from fractions import Fraction
from pathlib import Path
from lxml import etree
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

VERSION='1.2.0'
BASE_URL='https://webgeoas/ALKISBuch/index.html?fs='
A='http://www.adv-online.de/namespaces/adv/gid/7.1'; G='http://www.opengis.net/gml/3.2'
W='http://www.opengis.net/wfs/2.0'; X='http://www.w3.org/1999/xlink'
NS={'adv':A,'gml':G,'wfs':W,'xlink':X}; ID=f'{{{G}}}id'; HREF=f'{{{X}}}href'; TITLE=f'{{{X}}}title'
class WorkflowError(RuntimeError): pass

def text(v):
    if v is None:return ''
    if isinstance(v,float) and v.is_integer():return str(int(v))
    return str(v).strip()
def q(ns,t):return f'{{{ns}}}{t}'
def add(p,t,v=None,ns=A):
    e=etree.SubElement(p,q(ns,t))
    if v not in(None,''):e.text=text(v)
    return e
def sid(kind,*parts):
    a='0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz'; raw=(kind+'\x1f'+'\x1f'.join(text(x) for x in parts)).encode();n=int.from_bytes(hashlib.sha256(raw).digest()[:9],'big');z=''
    for _ in range(10):n,r=divmod(n,62);z=a[r]+z
    return 'DENIAL43'+z
def common(o,gid,stamp,event='000000',title='Ersteinrichtung',event_on=True):
    o.set(ID,gid);e=add(o,'identifier','urn:adv:oid:'+gid,G);e.set('codeSpace','http://www.adv-online.de/')
    e=add(o,'lebenszeitintervall');x=add(e,'AA_Lebenszeitintervall');add(x,'beginnt',stamp)
    e=add(o,'modellart');x=add(e,'AA_Modellart');add(x,'advStandardModell','DLKM')
    if event_on:e=add(o,'anlass');e.set(HREF,'https://registry.gdi-de.org/codelist/de.adv-online.gid/AA_Anlassart/'+event);e.set(TITLE,title)
def member(o):m=etree.Element(q(W,'member'));m.append(o);return m
def frac(v):
    m=re.fullmatch(r'\s*(-?\d+(?:[.,]\d+)?)\s*/\s*(-?\d+(?:[.,]\d+)?)\s*',text(v))
    if not m:return None
    try:z=Decimal(m.group(1).replace(',','.'));n=Decimal(m.group(2).replace(',','.'))
    except InvalidOperation:return None
    return None if n==0 else (z,n)
def decimal_text(v):
    if v==v.to_integral_value():return format(v,'.1f')
    return format(v,'f').rstrip('0').rstrip('.')
def add_share(o,f):
    e=add(o,'anteil');x=add(e,'AX_Anteil');add(x,'zaehler',decimal_text(f[0]));add(x,'nenner',decimal_text(f[1]))
def as_fraction(f):return Fraction(f[0])/Fraction(f[1])
def iso_date(v):
    z=text(v)
    for f in('%d.%m.%Y','%Y-%m-%d'):
        try:return datetime.strptime(z,f).strftime('%Y-%m-%d')
        except ValueError:pass
    return z
def person_key(r):return tuple(text(r.get(k)) for k in('Nachname oder Firma','Vorname','geb.','Geburtsdatum','Akademischer Grad','Anrede-Schlüssel','PLZ','Ort','Straße','Hausnummer','PLZ Postfach','Postfach','Land'))
def has_address(r):return any(text(r.get(k)) for k in('PLZ','Ort','Straße','Hausnummer','PLZ Postfach','Postfach','Land'))
def sheet_key(r):return text(r.get('Grundbuchblatt'))
def booking_key(r):return(sheet_key(r),text(r.get('Laufende Nummer')),text(r.get('Buchungsart')))
def plan(v):
    z=text(v);m=re.search(r'(\d+[A-Za-z]?)\s*$',z);return m.group(1) if m else z

def safe_xml(source,work):
    source=Path(source).resolve()
    if source.suffix.lower()=='.xml':return source,source.name
    if source.suffix.lower()!='.zip':raise WorkflowError('Bitte ZIP oder XML auswählen.')
    try:
        with zipfile.ZipFile(source) as z:
            xs=[i for i in z.infolist() if not i.is_dir() and i.filename.lower().endswith('.xml')]
            if len(xs)!=1:raise WorkflowError(f'Die ZIP muss genau eine XML enthalten. Gefunden: {len(xs)}.')
            target=Path(work)/Path(xs[0].filename).name
            with z.open(xs[0]) as a,target.open('wb') as b:shutil.copyfileobj(a,b)
            return target,xs[0].filename
    except zipfile.BadZipFile as e:raise WorkflowError('Beschädigte oder ungültige ZIP-Datei.') from e

def parse_parcels(xml):
    t=etree.parse(str(xml),etree.XMLParser(huge_tree=True,resolve_entities=False,no_network=True));out=[];seen=set()
    for e in t.xpath('//adv:AX_Flurstueck',namespaces=NS):
        oid=e.get(ID,'').strip()
        if not oid or oid in seen:continue
        seen.add(oid);z=e.xpath('string(adv:zaehler)',namespaces=NS).strip();n=e.xpath('string(adv:nenner)',namespaces=NS).strip()
        out.append({'DENIAL':oid,'Flurstueckskennzeichen':e.xpath('string(adv:flurstueckskennzeichen)',namespaces=NS).strip(),'Gemarkung':e.xpath('string(adv:gemarkung/adv:AX_Gemarkung_Schluessel/adv:gemarkungsnummer)',namespaces=NS).strip(),'Flur':e.xpath('string(adv:flurnummer)',namespaces=NS).strip(),'Flurstueck':z+('/'+n if n else '')})
    if not out:raise WorkflowError('Keine AX_Flurstueck-Objekte gefunden.')
    return out

def geoas_url(ids):return BASE_URL+"',".join(ids)
def write_url(path,url):Path(path).write_text('[InternetShortcut]\r\nURL='+url+'\r\n',encoding='utf-8-sig')
def write_html(path,url):
    safe=html.escape(url,quote=True);js=json.dumps(url,ensure_ascii=False)
    Path(path).write_text(f'''<!doctype html><html lang="de"><head><meta charset="utf-8"><title>GeoAS öffnen</title></head><body><h1>GeoAS – ALKIS IP Buch</h1><p>GeoAS wird mit den ausgewählten Flurstücken geöffnet.</p><p><a href="{safe}">Falls die Weiterleitung nicht startet: GeoAS hier öffnen</a></p><script>window.location.replace({js});</script></body></html>''',encoding='utf-8')
def open_url(path,url):os.startfile(str(path)) if os.name=='nt' else webbrowser.open(Path(path).resolve().as_uri())

def read_geoas(path):
    wb=load_workbook(path,data_only=True,read_only=True)
    try:
        ws=wb.active;it=ws.iter_rows(values_only=True)
        try:h=[text(x) for x in next(it)]
        except StopIteration:raise WorkflowError('Die GeoAS-XLSX ist leer.')
        required={'ALKIS_ObjektID','Grundbuchblatt','Buchungsart','Laufende Nummer','Namensnummer','Nachname oder Firma'};missing=required-set(h)
        if missing:raise WorkflowError('Erforderliche GeoAS-Spalten fehlen: '+', '.join(sorted(missing)))
        rows=[]
        for v in it:
            r=dict(zip(h,v))
            if text(r.get('ALKIS_ObjektID')):rows.append(r)
        return rows
    finally:
        wb.close()

def write_check(path,parcels,rows=None):
    rows=rows or [];gids={text(r.get('ALKIS_ObjektID')) for r in rows};info=defaultdict(lambda:{'codes':set(),'sheets':set(),'bvnrs':set()})
    for r in rows:
        oid=text(r.get('ALKIS_ObjektID'))
        if not oid:continue
        for c,k in [('Buchungsart','codes'),('Grundbuchblatt','sheets'),('Laufende Nummer','bvnrs')]:
            if text(r.get(c)):info[oid][k].add(text(r.get(c)))
    wb=Workbook();ws=wb.active;ws.title='DENIAL-Prüfung';ws.sheet_view.showGridLines=False
    hs=['Lfd. Nr.','DENIAL','Flurstückskennzeichen','Gemarkung','Flur','Flurstück','In OpenData','In GeoAS','Buchungsart','Grundbuchblatt','BVNR','Status','Hinweis'];ws.append(hs)
    for i,p in enumerate(parcels,1):
        found=p['DENIAL'] in gids if rows else None;b=info[p['DENIAL']];status='OK' if found else('FEHLT' if rows else'OFFEN')
        ws.append([i,p['DENIAL'],p['Flurstueckskennzeichen'],p['Gemarkung'],p['Flur'],p['Flurstueck'],'Ja','Ja' if found else('Nein' if rows else''),', '.join(sorted(b['codes'])),', '.join(sorted(b['sheets'])),', '.join(sorted(b['bvnrs'])),status,'' if status!='FEHLT' else'Keine Buchungszeile in GeoAS'])
    for c in ws[1]:c.fill=PatternFill('solid',fgColor='1F4E78');c.font=Font(color='FFFFFF',bold=True);c.alignment=Alignment(horizontal='center')
    widths=[10,24,28,12,8,14,14,12,18,24,12,12,42]
    for i,w in enumerate(widths,1):ws.column_dimensions[chr(64+i)].width=w
    ws.freeze_panes='A2';ws.auto_filter.ref=ws.dimensions
    colors={'OK':'E2F0D9','FEHLT':'F4CCCC','OFFEN':'FCE4D6'}
    for r in range(2,ws.max_row+1):ws.cell(r,12).fill=PatternFill('solid',fgColor=colors[ws.cell(r,12).value])
    ov=wb.create_sheet('Übersicht');ov.sheet_view.showGridLines=False;openids={p['DENIAL'] for p in parcels};matched=openids&gids;miss=openids-gids if rows else set()
    for row in [('ALKIS – GeoAS – GEOgraf','Wert'),('Programmversion',VERSION),('Flurstücke OpenData',len(openids)),('Flurstücke GeoAS',len(gids) if rows else'noch nicht geladen'),('Übereinstimmend',len(matched) if rows else'offen'),('Fehlend in GeoAS',len(miss) if rows else'offen')]:ov.append(row)
    for c in ov[1]:c.fill=PatternFill('solid',fgColor='1F4E78');c.font=Font(color='FFFFFF',bold=True)
    ov.column_dimensions['A'].width=34;ov.column_dimensions['B'].width=24;wb.save(path)

def prepare(source,project,auto_open=True):
    project=Path(project);project.mkdir(parents=True,exist_ok=True);temp=Path(tempfile.mkdtemp(prefix='alkis_',dir=project))
    try:
        xml,inner=safe_xml(source,temp);parcels=parse_parcels(xml);ids=[p['DENIAL'] for p in parcels];url=geoas_url(ids)
        (project/'01_DENIAL_Objekt-IDs.txt').write_text('\n'.join(ids)+'\n',encoding='utf-8');write_url(project/'02_GeoAS_Objekt-IDs.url',url);write_html(project/'02_GeoAS_Objekt-IDs.html',url);write_check(project/'03_DENIAL_Pruefung.xlsx',parcels)
        shutil.copy2(xml,project/'OpenData_NAS.xml');(project/'06_Verarbeitung.log').write_text(f'Version: {VERSION}\nQuelle: {source}\nXML: {inner}\nFlurstücke: {len(ids)}\nURL-Zeichen: {len(url)}\n',encoding='utf-8')
        if auto_open:open_url(project/'02_GeoAS_Objekt-IDs.html',url)
        return len(ids),url
    finally:shutil.rmtree(temp,ignore_errors=True)

def validate_xlsx(project,xlsx):
    project=Path(project);parcels=parse_parcels(project/'OpenData_NAS.xml');rows=read_geoas(xlsx);openids={p['DENIAL'] for p in parcels};geoids={text(r.get('ALKIS_ObjektID')) for r in rows};missing=openids-geoids;extra=geoids-openids
    write_check(project/'03_DENIAL_Pruefung.xlsx',parcels,rows)
    if extra:raise WorkflowError(f'{len(extra)} GeoAS-DENIAL(s) sind nicht in der OpenData-NAS enthalten.')
    shutil.copy2(xlsx,project/'GeoAS_Export.xlsx');mf=project/'07_Fehlende_GeoAS_Objekt-IDs.txt'
    if missing:mf.write_text('\n'.join(sorted(missing))+'\n',encoding='utf-8')
    elif mf.exists():mf.unlink()
    return {'opendata':len(openids),'geoas':len(geoids),'matched':len(openids&geoids),'missing':len(missing),'extra':0}

def make_catalog(rows,stamp):
    out=[];offices={};districts={}
    for r in rows:offices[text(r.get('Grundbuchamt-Schlüssel'))]=(text(r.get('Grundbuchamt-Name')),r);districts[text(r.get('Grundbuchbezirk-Schlüssel'))]=(text(r.get('Grundbuchbezirk-Name')),r)
    for key,(name,r) in offices.items():
        if not key:continue
        o=etree.Element(q(A,'AX_Dienststelle'));common(o,sid('DIENST',key),stamp);add(o,'schluesselGesamt',key);add(o,'bezeichnung',name);e=add(o,'schluessel');x=add(e,'AX_Dienststelle_Schluessel');add(x,'land',key[:2]);add(x,'stelle',key[2:]);add(o,'stellenart','1000');out.append(member(o))
    for key,(name,r) in districts.items():
        if not key:continue
        office=text(r.get('Grundbuchamt-Schlüssel'));o=etree.Element(q(A,'AX_Buchungsblattbezirk'));common(o,sid('BEZIRK',key),stamp);add(o,'schluesselGesamt',key);add(o,'bezeichnung',name);e=add(o,'schluessel');x=add(e,'AX_Buchungsblattbezirk_Schluessel');add(x,'land',key[:2]);add(x,'bezirk',key[2:])
        if office:e=add(o,'gehoertZu');x=add(e,'AX_Dienststelle_Schluessel');add(x,'land',office[:2]);add(x,'stelle',office[2:])
        out.append(member(o))
    return out

def generate_nas(nas,xlsx,out,report):
    tree=etree.parse(str(nas),etree.XMLParser(huge_tree=True,remove_blank_text=True,resolve_entities=False,no_network=True));fc=tree.xpath('//wfs:FeatureCollection',namespaces=NS)[0];rows=[r for r in read_geoas(xlsx) if sheet_key(r)];by=defaultdict(list)
    for r in rows:by[text(r['ALKIS_ObjektID'])].append(r)
    stamp=datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace('+00:00','Z');made=[];warn=[];stats=Counter();classes=Counter();sheets={};positions={};people={};addresses={}
    for r in rows:
        if not text(r.get('Nachname oder Firma')):continue
        k=person_key(r)
        if k in people:continue
        people[k]=sid('PERSON',*k);o=etree.Element(q(A,'AX_Person'));common(o,people[k],stamp,'050000','Angaben zu Eigentümer oder Erbbauberechtigten verändern');add(o,'nachnameOderFirma',r.get('Nachname oder Firma'));add(o,'anrede',r.get('Anrede-Schlüssel'));add(o,'vorname',r.get('Vorname'));add(o,'akademischerGrad',r.get('Akademischer Grad'));add(o,'geburtsname',r.get('geb.'));add(o,'geburtsdatum',iso_date(r.get('Geburtsdatum')))
        if has_address(r):addresses[k]=sid('ANSCHRIFT',*k);e=add(o,'hat');e.set(HREF,'urn:adv:oid:'+addresses[k])
        made.append(member(o));stats['AX_Person']+=1
    for k,aid in addresses.items():
        r=next(x for x in rows if text(x.get('Nachname oder Firma')) and person_key(x)==k);o=etree.Element(q(A,'AX_Anschrift'));common(o,aid,stamp,event_on=False);add(o,'ort_Post',r.get('Ort'));add(o,'postleitzahlPostzustellung',r.get('PLZ'));add(o,'strasse',r.get('Straße'));add(o,'hausnummer',r.get('Hausnummer'));add(o,'postleitzahlPostfach',r.get('PLZ Postfach'));add(o,'postfach',r.get('Postfach'));add(o,'bestimmungsland',r.get('Land'));made.append(member(o));stats['AX_Anschrift']+=1
    for r in rows:
        sk=sheet_key(r)
        if sk not in sheets:
            sheets[sk]=sid('BLATT',sk);o=etree.Element(q(A,'AX_Buchungsblatt'));common(o,sheets[sk],stamp);district=text(r.get('Grundbuchbezirk-Schlüssel'));add(o,'buchungsblattkennzeichen',re.sub(r'[^0-9A-Za-z]','',sk));e=add(o,'buchungsblattbezirk');x=add(e,'AX_Buchungsblattbezirk_Schluessel');add(x,'land',district[:2]);add(x,'bezirk',district[2:]);add(o,'buchungsblattnummerMitBuchstabenerweiterung',r.get('Grundbuchblattnummer'));add(o,'blattart',r.get('Buchungsblattart'));made.append(member(o));stats['AX_Buchungsblatt']+=1
        positions.setdefault(booking_key(r),sid('STELLE',*booking_key(r)))
    special={}
    for oid,rr in by.items():
        codes={text(x.get('Buchungsart')) for x in rr}
        if codes in ({'1301'},{'1302'}):
            code=next(iter(codes));key=(code,tuple(sorted({booking_key(x) for x in rr})));special.setdefault(key,[]).append(oid)
    fake=None
    if special:
        sm=next(iter(by.values()))[0];d=text(sm.get('Grundbuchbezirk-Schlüssel'));fake=sid('FIKTIV-AUFTEILUNG','9999999');o=etree.Element(q(A,'AX_Buchungsblatt'));common(o,fake,stamp);add(o,'buchungsblattkennzeichen',d+'9999999');e=add(o,'buchungsblattbezirk');x=add(e,'AX_Buchungsblattbezirk_Schluessel');add(x,'land',d[:2]);add(x,'bezirk',d[2:]);add(o,'buchungsblattnummerMitBuchstabenerweiterung','9999999');add(o,'blattart','5000');made.append(member(o));stats['AX_Buchungsblatt']+=1
    direct={};emitted=set()
    for idx,(key,oids) in enumerate(sorted(special.items(),key=lambda x:x[0]),1):
        code,bks=key;root=sid('AUFTEILUNGS-WURZEL',code,*map(str,bks))
        for oid in oids:direct[oid]=root
        o=etree.Element(q(A,'AX_Buchungsstelle'));common(o,root,stamp);add(o,'buchungsart','1101' if code=='1301' else '1102');add(o,'laufendeNummer',str(idx).zfill(4));e=add(o,'istBestandteilVon');e.set(HREF,'urn:adv:oid:'+fake);made.append(member(o));stats['AX_Buchungsstelle']+=1;stats['Komplex_'+code]+=1;total=Fraction();rr=by[oids[0]]
        for bk in bks:
            if bk in emitted:continue
            emitted.add(bk);r=next(x for x in rr if booking_key(x)==bk);o=etree.Element(q(A,'AX_Buchungsstelle'));common(o,positions[bk],stamp);add(o,'buchungsart',code);add(o,'laufendeNummer',text(r.get('Laufende Nummer')).zfill(4));f=frac(r.get('Miteigentumsanteil'))
            if f:add_share(o,f);total+=as_fraction(f);stats['Dezimalanteile']+=int(',' in text(r.get('Miteigentumsanteil')) or '.' in text(r.get('Miteigentumsanteil')).split('/')[0])
            else:warn.append('/'.join(oids)+': ungültiger Anteil')
            if code=='1301':add(o,'nummerImAufteilungsplan',plan(r.get('Nummer im Aufteilungsplan')))
            e=add(o,'istBestandteilVon');e.set(HREF,'urn:adv:oid:'+sheets[bk[0]]);e=add(o,'an');e.set(HREF,'urn:adv:oid:'+root);made.append(member(o));stats['AX_Buchungsstelle']+=1
        if total!=1:warn.append(code+'-Komplex '+('/'.join(oids))+f': Summe {total}, erwartet 1')
    for oid,rr in sorted(by.items()):
        codes={text(r.get('Buchungsart')) for r in rr};classes[tuple(sorted(codes))]+=1
        if codes in ({'1301'},{'1302'}):continue
        primary=[r for r in rr if text(r.get('Buchungsart'))=='1100'] or[rr[0]];pbk=booking_key(primary[0]);direct[oid]=positions[pbk]
        for bk in sorted({booking_key(r) for r in rr}):
            if bk in emitted:continue
            emitted.add(bk);r=next(x for x in rr if booking_key(x)==bk);code=text(r.get('Buchungsart'));o=etree.Element(q(A,'AX_Buchungsstelle'));common(o,positions[bk],stamp,'070000' if code=='2101' else'000000','Rechte buchen' if code=='2101' else'Ersteinrichtung');add(o,'buchungsart',code);add(o,'laufendeNummer',text(r.get('Laufende Nummer')).zfill(4));e=add(o,'istBestandteilVon');e.set(HREF,'urn:adv:oid:'+sheets[bk[0]])
            if code=='2101':e=add(o,'an');e.set(HREF,'urn:adv:oid:'+positions[pbk])
            made.append(member(o));stats['AX_Buchungsstelle']+=1
    seen=set()
    for r in rows:
        sk=sheet_key(r);nn=text(r.get('Namensnummer'));isp=bool(text(r.get('Nachname oder Firma')));comm=bool(text(r.get('Art der Rechtsgemeinschaft')) or text(r.get('Beschrieb der Rechtsgemeinschaft')))
        if not nn or(not isp and not comm):continue
        nk=(sk,nn,person_key(r) if isp else('COMMUNITY',text(r.get('Beschrieb der Rechtsgemeinschaft'))))
        if nk in seen:continue
        seen.add(nk);o=etree.Element(q(A,'AX_Namensnummer'));common(o,sid('NAMEN',*map(str,nk)),stamp,'050000','Angaben zu Eigentümer oder Erbbauberechtigten verändern')
        if isp:e=add(o,'benennt');e.set(HREF,'urn:adv:oid:'+people[person_key(r)])
        add(o,'laufendeNummerNachDIN1421',nn);f=frac(r.get('Gemeinschaftl. Anteil'))
        if f:add_share(o,f)
        if comm:art=text(r.get('Art der Rechtsgemeinschaft'));add(o,'artDerRechtsgemeinschaft','9999' if art.lower()=='sonstiges' else art);add(o,'beschriebDerRechtsgemeinschaft',r.get('Beschrieb der Rechtsgemeinschaft'));stats['Rechtsgemeinschaft']+=1
        e=add(o,'istBestandteilVon');e.set(HREF,'urn:adv:oid:'+sheets[sk]);made.append(member(o));stats['AX_Namensnummer']+=1
    cats=make_catalog(rows,stamp);made.extend(cats);stats['Katalogobjekte']=len(cats);members=fc.findall(q(W,'member'));idx=next(i for i,m in enumerate(members) if len(m) and etree.QName(m[0]).localname=='AX_Flurstueck')
    for j,m in enumerate(made):fc.insert(idx+j,m)
    missing=[]
    for oid,target in direct.items():
        found=tree.xpath('//adv:AX_Flurstueck[@gml:id=$i]',namespaces=NS,i=oid)
        if not found:missing.append(oid);continue
        p=found[0]
        for old in list(p.findall(q(A,'istGebucht'))):p.remove(old)
        pos=next((i for i,c in enumerate(p) if etree.QName(c).localname in('weistAuf','zeigtAuf')),len(p));e=etree.Element(q(A,'istGebucht'));e.set(HREF,'urn:adv:oid:'+target);p.insert(pos,e)
    n=len(fc.findall(q(W,'member')));fc.set('numberMatched',str(n));fc.set('numberReturned',str(n));etree.indent(tree.getroot(),space='');tree.write(str(out),encoding='UTF-8',xml_declaration=True,pretty_print=True)
    check=etree.parse(str(out),etree.XMLParser(huge_tree=True));ids=[e.get(ID) for e in check.xpath('//*[@gml:id]',namespaces=NS)];idset=set(ids);types=('AX_Buchungsstelle','AX_Buchungsblatt','AX_Namensnummer','AX_Person','AX_Anschrift','AX_Dienststelle','AX_Buchungsblattbezirk');sources=check.xpath(' | '.join('//adv:'+x for x in types),namespaces=NS)+[e for oid in direct for e in check.xpath('//adv:AX_Flurstueck[@gml:id=$i]',namespaces=NS,i=oid)];refs=[h for e in sources for h in e.xpath('.//@xlink:href',namespaces=NS) if h.startswith('urn:adv:oid:')];broken=[h for h in refs if h[12:] not in idset]
    lines=['GEOgraf NAS-Assistent – Prüfprotokoll',f'Version: {VERSION}',f'Zeitpunkt: {stamp}',f'GeoAS-Zeilen mit Buchung: {len(rows)}',f'Mit GeoAS verarbeitete Flurstücke: {len(by)}',f'GeoAS-Flurstücke nicht in NAS: {len(missing)}','', 'Fallgruppen:']+[f'  {" + ".join(k)}: {v}' for k,v in sorted(classes.items())]+['','Erzeugte Objekte:']+[f'  {k}: {v}' for k,v in sorted(stats.items())]+['',f'Doppelte gml:id: {sum(v>1 for v in Counter(ids).values())}',f'Gebrochene Referenzen in erzeugten Ketten: {len(broken)}',f'Warnungen: {len(warn)}']+(['  '+w for w in warn] if warn else['  keine']);Path(report).write_text('\n'.join(lines)+'\n',encoding='utf-8')
    if missing or broken or any(v>1 for v in Counter(ids).values()):raise WorkflowError('Integritätsprüfung fehlgeschlagen; siehe 05_Pruefprotokoll.txt.')

def generate_project(project):
    project=Path(project);out=project/'04_GEOgraf_NAS_mit_Eigentuemer.xml';generate_nas(project/'OpenData_NAS.xml',project/'GeoAS_Export.xlsx',out,project/'05_Pruefprotokoll.txt');return out

def gui():
    import tkinter as tk
    from tkinter import filedialog,messagebox,ttk
    root=tk.Tk();root.title('OpenGeoData.NI → GeoAS → GEOgraf');root.geometry('780x545');root.minsize(720,510);src=tk.StringVar();xls=tk.StringVar();proj=tk.StringVar();status=tk.StringVar(value='Bereit – bitte OpenData-ZIP oder XML auswählen.');body=ttk.Frame(root,padding=18);body.pack(fill='both',expand=True)
    ttk.Label(body,text='ALKIS – GeoAS – GEOgraf Assistent',font=('Segoe UI',18,'bold')).grid(row=0,column=0,columnspan=3,sticky='w',pady=(0,18));ttk.Label(body,text='1. OpenData').grid(row=1,column=0,sticky='w');ttk.Entry(body,textvariable=src,width=74).grid(row=2,column=0,columnspan=2,sticky='ew',padx=(0,8))
    def choose_src():
        f=filedialog.askopenfilename(filetypes=[('OpenData ZIP/XML','*.zip *.xml'),('Alle Dateien','*.*')])
        if f:src.set(f);proj.set(str(Path(f).with_suffix(''))+'_GEOgraf')
    ttk.Button(body,text='ZIP/XML auswählen',command=choose_src).grid(row=2,column=2);ttk.Label(body,text='Projektordner').grid(row=3,column=0,sticky='w',pady=(10,0));ttk.Entry(body,textvariable=proj,width=74).grid(row=4,column=0,columnspan=2,sticky='ew',padx=(0,8))
    box=tk.Label(body,textvariable=status,anchor='w',justify='left',bg='#E2F0D9',padx=12,pady=12,font=('Segoe UI',10,'bold'))
    def run_prepare():
        try:n,_=prepare(src.get(),proj.get(),True);box.configure(bg='#E2F0D9');status.set(f'GRÜN – {n} Flurstücke; GeoAS wurde automatisch geöffnet.')
        except Exception as e:box.configure(bg='#F4CCCC');status.set('ROT – '+str(e));messagebox.showerror('Fehler',str(e))
    ttk.Button(body,text='DENIALs erzeugen + GeoAS öffnen',command=run_prepare).grid(row=4,column=2);ttk.Separator(body).grid(row=5,column=0,columnspan=3,sticky='ew',pady=18);ttk.Label(body,text='2. GeoAS-Export').grid(row=6,column=0,sticky='w');ttk.Entry(body,textvariable=xls,width=74).grid(row=7,column=0,columnspan=2,sticky='ew',padx=(0,8))
    def choose_xls():
        f=filedialog.askopenfilename(filetypes=[('Excel-Arbeitsmappe','*.xlsx')])
        if f:xls.set(f)
    ttk.Button(body,text='XLSX auswählen',command=choose_xls).grid(row=7,column=2)
    def run_generate():
        try:
            result=validate_xlsx(proj.get(),xls.get());out=generate_project(proj.get())
            if result['missing']:
                box.configure(bg='#FCE4D6');status.set(f'GELB – {result["matched"]} von {result["opendata"]} zugeordnet; {result["missing"]} fehlen. NAS wurde trotzdem erzeugt.');messagebox.showwarning('NAS mit unvollständigen GeoAS-Daten erzeugt',f'Die NAS wurde erzeugt.\n\nMit GeoAS-Daten: {result["matched"]}\nOhne GeoAS-Daten: {result["missing"]}\n\nDetails: 03_DENIAL_Pruefung.xlsx und 07_Fehlende_GeoAS_Objekt-IDs.txt')
            else:
                box.configure(bg='#E2F0D9');status.set(f'GRÜN – {result["matched"]} von {result["opendata"]} zugeordnet; NAS fertig.');messagebox.showinfo('Fertig',f'GEOgraf-NAS erzeugt:\n{out}')
        except Exception as e:box.configure(bg='#F4CCCC');status.set('ROT – '+str(e));messagebox.showerror('Fehler',str(e))
    ttk.Button(body,text='GEOgraf-NAS erzeugen',command=run_generate).grid(row=9,column=0,pady=22,sticky='w')
    def open_dir():
        p=Path(proj.get());p.mkdir(parents=True,exist_ok=True);os.startfile(str(p)) if os.name=='nt' else webbrowser.open(p.as_uri())
    ttk.Button(body,text='Ausgabeordner öffnen',command=open_dir).grid(row=9,column=1,pady=22,sticky='w')
    box.grid(row=10,column=0,columnspan=3,sticky='ew')
    footer=ttk.Label(
        body,
        text=f'OE 61.14, Daniel Bernutz | Version {VERSION} | Stand: 15.07.2026',
        font=('Segoe UI',8),
        foreground='#666666',
        anchor='e'
    )
    footer.grid(row=11,column=0,columnspan=3,sticky='e',pady=(12,0))
    body.columnconfigure(0,weight=1);body.columnconfigure(1,weight=1);root.mainloop()

def main():
    ap=argparse.ArgumentParser();ap.add_argument('--prepare');ap.add_argument('--project');ap.add_argument('--xlsx');ap.add_argument('--generate',action='store_true');ap.add_argument('--no-open',action='store_true');a=ap.parse_args()
    if not any((a.prepare,a.xlsx,a.generate)):return gui()
    if not a.project:raise SystemExit('--project ist erforderlich')
    if a.prepare:print(prepare(a.prepare,a.project,not a.no_open))
    if a.xlsx:print(validate_xlsx(a.project,a.xlsx))
    if a.generate:print(generate_project(a.project))
if __name__=='__main__':
    try:main()
    except Exception as e:print('FEHLER:',e,file=sys.stderr);traceback.print_exc();raise SystemExit(1)
